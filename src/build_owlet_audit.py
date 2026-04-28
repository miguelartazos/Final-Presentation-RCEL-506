from __future__ import annotations

import warnings
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
import ternary
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parent / "outputs" / "owlet-audit"
OUT_DIR.mkdir(parents=True, exist_ok=True)
warnings.filterwarnings("ignore", message="No data for colormapping provided via 'c'")

XLSX_PATH = OUT_DIR / "owlet_managerial_energy_log.xlsx"
PNG_PATH = OUT_DIR / "owlet_managerial_balance_ternary.png"
CSV_PATH = OUT_DIR / "owlet_managerial_energy_log.csv"
SUMMARY_PATH = OUT_DIR / "owlet_hiring_decision.txt"
PDF_PATH = OUT_DIR / "owlet_performance_evaluation.pdf"


INCIDENTS = [
    {
        "Incident #": 1,
        "Critical Incident": "I had to redefine what 'balance' actually meant in the project so fairness was measured in a clear way instead of with a vague custom score.",
        "Failure Type": "Intent Misalignment",
        "Governance": 60,
        "Audit": 20,
        "Context": 20,
    },
    {
        "Incident #": 2,
        "Critical Incident": "I had to correct the idea that simulation and machine learning were basically the same thing and explain the role of each one more clearly.",
        "Failure Type": "Domain Misunderstanding",
        "Governance": 55,
        "Audit": 25,
        "Context": 20,
    },
    {
        "Incident #": 3,
        "Critical Incident": "I had to stop OWLET from pushing a much more advanced graph-AI solution that sounded impressive but did not fit the size of my project.",
        "Failure Type": "Overengineering / Weakly-Informed",
        "Governance": 65,
        "Audit": 10,
        "Context": 25,
    },
    {
        "Incident #": 4,
        "Critical Incident": "Stopped the repo from depending on an external workflow editor and redirected the presentation diagrams into reproducible in-repo PNG assets.",
        "Failure Type": "Intent Misalignment",
        "Governance": 50,
        "Audit": 15,
        "Context": 35,
    },
    {
        "Incident #": 5,
        "Critical Incident": "Caught number drift between old midterm deck counts and the live parser output, preventing older 131-card figures from contaminating the final story.",
        "Failure Type": "Hallucination / Stale Facts",
        "Governance": 20,
        "Audit": 55,
        "Context": 25,
    },
    {
        "Incident #": 6,
        "Critical Incident": "Rewrote the ethics-audit example so it used a real project near-miss instead of a polished but partly invented 'classist auto-balancer' scenario.",
        "Failure Type": "Hallucination",
        "Governance": 35,
        "Audit": 45,
        "Context": 20,
    },
    {
        "Incident #": 7,
        "Critical Incident": "I had to fix the optimization workflow after it started flagging its own internal files as if they were errors.",
        "Failure Type": "Domain / Tooling Misunderstanding",
        "Governance": 50,
        "Audit": 35,
        "Context": 15,
    },
    {
        "Incident #": 8,
        "Critical Incident": "I found out that the optimization tool was tracking the wrong result because the output format was incomplete.",
        "Failure Type": "Hallucination / Wrong Metric",
        "Governance": 45,
        "Audit": 45,
        "Context": 10,
    },
    {
        "Incident #": 9,
        "Critical Incident": "Noticed that the presentation notes still used outdated strategy names, card counts, and business claims from the midterm build.",
        "Failure Type": "Context Decay",
        "Governance": 25,
        "Audit": 45,
        "Context": 30,
    },
    {
        "Incident #": 10,
        "Critical Incident": "I had to refocus the final deck so it was not just about charts, but about showing a real data-science process with evidence and business use.",
        "Failure Type": "Intent Misalignment",
        "Governance": 65,
        "Audit": 5,
        "Context": 30,
    },
]


def build_workbook() -> pd.DataFrame:
    df = pd.DataFrame(INCIDENTS)
    df["Total"] = df[["Governance", "Audit", "Context"]].sum(axis=1)
    df.to_csv(CSV_PATH, index=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Energy Log"

    headers = [
        "Incident #",
        "Critical Incident",
        "Failure Type",
        "Governance",
        "Audit",
        "Context",
        "Total",
    ]
    ws.append(headers)

    for row_idx, record in enumerate(INCIDENTS, start=2):
        ws.cell(row=row_idx, column=1, value=record["Incident #"])
        ws.cell(row=row_idx, column=2, value=record["Critical Incident"])
        ws.cell(row=row_idx, column=3, value=record["Failure Type"])
        ws.cell(row=row_idx, column=4, value=record["Governance"])
        ws.cell(row=row_idx, column=5, value=record["Audit"])
        ws.cell(row=row_idx, column=6, value=record["Context"])
        ws.cell(row=row_idx, column=7, value=f"=SUM(D{row_idx}:F{row_idx})")

    avg_row = len(INCIDENTS) + 2
    ws.cell(row=avg_row, column=1, value="Averages")
    ws.cell(row=avg_row, column=4, value=f"=AVERAGE(D2:D{avg_row-1})")
    ws.cell(row=avg_row, column=5, value=f"=AVERAGE(E2:E{avg_row-1})")
    ws.cell(row=avg_row, column=6, value=f"=AVERAGE(F2:F{avg_row-1})")
    ws.cell(row=avg_row, column=7, value=f"=SUM(D{avg_row}:F{avg_row})")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    avg_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[avg_row]:
        cell.fill = avg_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=avg_row - 1, min_col=4, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=avg_row - 1, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 12,
        2: 80,
        3: 32,
        4: 14,
        5: 12,
        6: 12,
        7: 10,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary["A1"] = "OWLET Performance Audit Summary"
    summary["A1"].font = Font(size=14, bold=True)
    summary["A3"] = "Average Governance"
    summary["B3"] = f"=ROUND('Energy Log'!D{avg_row},2)"
    summary["A4"] = "Average Audit"
    summary["B4"] = f"=ROUND('Energy Log'!E{avg_row},2)"
    summary["A5"] = "Average Context"
    summary["B5"] = f"=ROUND('Energy Log'!F{avg_row},2)"
    summary["A7"] = "Recommended Hiring Decision"
    summary["B7"] = "RETRAIN"
    summary["A9"] = "Interpretation"
    summary["B9"] = (
        "The strongest pull is toward Governance, which means most manager time was spent correcting domain framing, "
        "workflows, and project intent rather than simply accepting output."
    )
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 95
    for cell in ("A3", "A4", "A5", "A7", "A9"):
        summary[cell].font = bold_font
    summary["B9"].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(XLSX_PATH)
    return df


def build_plot(df: pd.DataFrame) -> tuple[float, float, float]:
    g_avg = round(df["Governance"].mean(), 2)
    a_avg = round(df["Audit"].mean(), 2)
    c_avg = round(df["Context"].mean(), 2)

    scale = 100
    figure, tax = ternary.figure(scale=scale)
    figure.set_size_inches(10, 8)

    tax.boundary(linewidth=2.0)
    tax.gridlines(color="#D9D9D9", multiple=10, linewidth=0.8)
    tax.scatter(
        [(g_avg, a_avg, c_avg)],
        marker="o",
        color="#1f77b4",
        s=120,
        zorder=5,
    )
    tax.annotate(
        f"Maiky\nG={g_avg} A={a_avg} C={c_avg}",
        (g_avg, a_avg, c_avg),
        fontsize=11,
        xytext=(0, 14),
        textcoords="offset points",
        ha="center",
    )
    tax.bottom_axis_label("Audit", fontsize=12, offset=0.14)
    tax.right_axis_label("Context", fontsize=12, offset=0.14)
    tax.left_axis_label("Governance", fontsize=12, offset=0.14)
    tax.ticks(axis="lbr", linewidth=1, multiple=10, fontsize=10, offset=0.02)
    tax.clear_matplotlib_ticks()
    tax.set_title("Project Managerial Balance Map", fontsize=16, pad=20)
    plt.tight_layout()
    plt.savefig(PNG_PATH, dpi=220, bbox_inches="tight")
    plt.close(figure)
    return g_avg, a_avg, c_avg


def write_reflection(g_avg: float, a_avg: float, c_avg: float) -> None:
    text = f"""Hiring Decision: RETRAIN

Average managerial balance:
- Governance: {g_avg}
- Audit: {a_avg}
- Context: {c_avg}

Why:
My data shows that the biggest source of management debt was Governance, not pure hallucination. That means OWLET was most useful once I constrained it, but it still needed regular correction on framing, domain logic, and workflow boundaries. The semester pattern was not sustainable enough to justify a full 'hire' decision, but it was also not bad enough to justify termination because many of the failures were fixable through better SOPs, benchmark contracts, and presentation rules.

What to retrain:
- A project brief that clearly separates simulation, optimization, and machine learning
- A metrics contract that forces baseline comparison and explicit score fields
- A presentation checklist that prevents stale counts, weak business framing, and outdated project notes

General conclusion:
My professional opinion is that GenAI behaves less like a finished employee and more like a fast trainee. It can produce useful drafts, code, and structure, but only under strong managerial oversight. The real engineering-management skill is not just delegating work to AI; it is knowing when to redirect, verify, and constrain it before a polished mistake becomes a project risk.
"""
    SUMMARY_PATH.write_text(text, encoding="utf-8")


def build_pdf_report(df: pd.DataFrame, g_avg: float, a_avg: float, c_avg: float) -> None:
    failure_counts = df["Failure Type"].value_counts()
    top_failure = str(failure_counts.index[0])
    top_failure_count = int(failure_counts.iloc[0])

    fig = plt.figure(figsize=(11, 8.5), facecolor="#F7EFDD")
    canvas = fig.add_axes([0, 0, 1, 1])
    canvas.axis("off")

    canvas.text(
        0.06,
        0.93,
        "OWLET Performance Evaluation",
        fontsize=24,
        fontweight="bold",
        color="#1D2824",
    )
    canvas.text(
        0.06,
        0.885,
        "Managerial oversight report for the Business Empire data science final",
        fontsize=11,
        color="#686658",
    )
    canvas.text(
        0.82,
        0.925,
        "Decision: RETRAIN",
        fontsize=13,
        fontweight="bold",
        color="#A94532",
        ha="center",
        bbox=dict(boxstyle="round,pad=0.45", facecolor="#FFF2D8", edgecolor="#D98B26"),
    )

    metrics = [
        ("Incidents audited", len(df), "#1E6D8C"),
        ("Top failure type", top_failure_count, "#A94532"),
        ("Governance avg", g_avg, "#2E5F4D"),
        ("Audit avg", a_avg, "#D98B26"),
    ]
    for idx, (label, value, color) in enumerate(metrics):
        x = 0.06 + idx * 0.225
        canvas.text(x, 0.80, str(value), fontsize=24, fontweight="bold", color=color)
        canvas.text(x, 0.765, label, fontsize=9.5, color="#686658")

    bar_ax = fig.add_axes([0.08, 0.46, 0.24, 0.22], facecolor="#F7EFDD")
    categories = ["Governance", "Audit", "Context"]
    values = [g_avg, a_avg, c_avg]
    colors = ["#2E5F4D", "#D98B26", "#1E6D8C"]
    bar_ax.bar(categories, values, color=colors)
    bar_ax.set_ylim(0, 75)
    bar_ax.set_title("Original metric: managerial energy mix", fontsize=11, fontweight="bold", pad=10)
    bar_ax.tick_params(axis="x", labelrotation=0, labelsize=9)
    bar_ax.tick_params(axis="y", labelsize=8)
    bar_ax.spines[["top", "right"]].set_visible(False)
    for idx, value in enumerate(values):
        bar_ax.text(idx, value + 2.0, f"{value:.1f}", ha="center", fontsize=9, fontweight="bold")

    failure_ax = fig.add_axes([0.62, 0.46, 0.33, 0.22], facecolor="#F7EFDD")
    failure_counts.sort_values().plot.barh(ax=failure_ax, color="#1E6D8C")
    failure_ax.set_title("Intervention pattern", fontsize=11, fontweight="bold", pad=10)
    failure_ax.set_ylabel("")
    failure_ax.tick_params(axis="both", labelsize=8)
    failure_ax.spines[["top", "right"]].set_visible(False)

    canvas.text(0.06, 0.36, "Strategic redirection example", fontsize=13, fontweight="bold", color="#1D2824")
    canvas.text(
        0.06,
        0.295,
        (
            "OWLET pushed a graph-AI direction that sounded advanced but did not fit the project scale.\n"
            "I redirected the work toward simulator-grounded balance metrics, baseline comparisons,\n"
            "and a code path that someone reviewing the project could inspect."
        ),
        fontsize=10.5,
        color="#1D2824",
        verticalalignment="top",
        linespacing=1.5,
    )

    canvas.text(0.06, 0.19, "Conclusion", fontsize=13, fontweight="bold", color="#1D2824")
    canvas.text(
        0.06,
        0.125,
        (
            "The AI was useful as a fast trainee, but not reliable as an unsupervised analyst.\n"
            "Most manager effort went into governance: correcting framing, constraining scope,\n"
            "and verifying that polished outputs matched the real project data."
        ),
        fontsize=10.5,
        color="#1D2824",
        verticalalignment="top",
        linespacing=1.5,
    )
    canvas.text(0.06, 0.04, f"Source data: {CSV_PATH.name} | Generated by build_owlet_audit.py", fontsize=8.5, color="#686658")

    fig.savefig(PDF_PATH, format="pdf", facecolor=fig.get_facecolor())
    plt.close(fig)


def main() -> None:
    df = build_workbook()
    g_avg, a_avg, c_avg = build_plot(df)
    write_reflection(g_avg, a_avg, c_avg)
    build_pdf_report(df, g_avg, a_avg, c_avg)
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved plot: {PNG_PATH}")
    print(f"Saved PDF report: {PDF_PATH}")
    print(f"Averages -> Governance={g_avg}, Audit={a_avg}, Context={c_avg}")


if __name__ == "__main__":
    main()
